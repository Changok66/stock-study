"""watch_trades.py가 새 체결을 감지할 때마다 호출되어, data/trade_log.csv를 기준으로
통합_매매일지_롤링서식.xlsx의 계좌 시트를 완전히 지우고 처음부터 다시 그린다.

라운드 분리 / 블록 작성 로직 자체는 build_dynamic_rounds.py에 있고, 이 파일은
CSV 읽기 -> 워크북 열기 -> 재작성 -> 저장을 이어주는 역할만 한다.
"""

import csv
import os
import sys

from dotenv import load_dotenv
from openpyxl import load_workbook

import build_dynamic_rounds as rounds_builder

load_dotenv()

LOG_FILE = "data/trade_log.csv"
XLSX_FILE = "data/통합_매매일지_롤링서식.xlsx"
ACCOUNT_SHEET_MAP = {
    os.getenv("KIWOOM_ACCOUNT_NO", ""): "국내1",
}


def _load_account_rows(account_no):
    if not os.path.exists(LOG_FILE):
        return []
    with open(LOG_FILE, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return [row for row in reader if row.get("계좌번호") == account_no]


def update_journal(account_no=None, sheet_name=None):
    account_no = account_no or os.getenv("KIWOOM_ACCOUNT_NO", "")
    sheet_name = sheet_name or ACCOUNT_SHEET_MAP.get(account_no, "국내1")

    rows = _load_account_rows(account_no)
    if not rows:
        print("[INFO] 매매일지 갱신 건너뜀: 해당 계좌의 체결 기록이 없습니다.")
        return

    all_rounds = rounds_builder.build_rounds(rows)

    try:
        wb = load_workbook(XLSX_FILE)
    except FileNotFoundError:
        print(f"[FAIL] 매매일지 엑셀 파일을 찾을 수 없습니다: {XLSX_FILE}")
        return

    if sheet_name not in wb.sheetnames:
        print(f"[FAIL] '{sheet_name}' 시트를 찾을 수 없습니다: {XLSX_FILE}")
        return

    ws = wb[sheet_name]
    rounds_builder.rebuild_sheet(ws, all_rounds)

    try:
        wb.save(XLSX_FILE)
    except PermissionError:
        print(f"[FAIL] 매매일지 저장 실패: {XLSX_FILE} 파일이 다른 프로그램(엑셀 등)에서 열려 있습니다.")
        return

    print(f"[완료] 매매일지 갱신: {sheet_name} 시트에 라운드 {len(all_rounds)}개 반영")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    update_journal()
