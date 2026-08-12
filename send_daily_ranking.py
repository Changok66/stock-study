"""
오늘 만든 3가지 키움 순위 조회 결과
(실시간종목조회순위 / 거래대금상위 / 시가총액상위)를 카테고리별로 별도의
카카오톡 "나에게 보내기" 메시지로 순차 전송하는 스크립트.

재사용 원칙:
- 코드를 새로 베껴 쓰지 않고, 이미 만들어둔 스크립트를 파이썬 모듈로 import해서
  그 안의 함수를 그대로 호출하는 방식으로 재사용한다.
  - 실시간종목조회순위(ka00198)/거래대금상위(ka10032) 조회: get_kiwoom_ranking.py
  - 시가총액상위(ka10099 기반) 조회: get_kiwoom_marketcap_ranking.py
  - 카카오톡 "나에게 보내기" 전송: send_kakao_message.py
- 세 스크립트 모두 `if __name__ == "__main__":` 가드로 감싸져 있어서,
  import만 해서는 자동으로 실행되지 않고 함수만 가져다 쓸 수 있다.

전체 흐름:
1. get_kiwoom_ranking.load_app_credentials()/issue_access_token()으로
   키움 접근토큰을 "한 번만" 발급받는다. (조회순위/거래대금/시가총액 조회 3번 모두
   이 토큰 하나를 공용으로 사용 - 매번 새로 발급받을 필요가 없다)
2. get_kiwoom_ranking.request_ranking()을 두 번 호출해
   실시간종목조회순위(ka00198), 거래대금상위(ka10032)를 각각 조회하고,
   ETF/ETN·우선주를 제외한 뒤 상위 TOP_N개만 추린다.
   (ka00198은 ka10030/ka10032와 달리 거래량/거래대금이 아니라 "실시간 조회 빈도"
    기준 순위이며, 등락률에 해당하는 필드명도 flu_rt가 아니라 base_comp_chgr다.
    get_kiwoom_ranking.request_ranking()이 TR_ID에 따라 요청 URL/payload를
    알아서 분기하므로 이 스크립트에서는 TR_ID만 바꿔주면 된다.)
3. get_kiwoom_marketcap_ranking.request_stock_list()로 KOSPI+KOSDAQ 전체
   종목을 모은 뒤, 같은 방식으로 ETF/ETN·우선주를 제외하고 시가총액
   (상장주식수 x 현재가) 기준으로 정렬해 상위 TOP_N개를 추린다.
4. 세 카테고리(실시간종목조회순위/거래대금상위/시가총액상위) TOP_N(종목명/종목코드/등락률
   또는 시가총액)을 data/daily_ranking_log.csv에 이어서(append) 기록한다. 파일이 없으면
   헤더를 포함해 새로 만들고, 있으면 헤더 없이 행만 추가한다. 이 기록은 카카오톡 전송
   성공 여부와 무관하게 조회할 때마다 남긴다.
5. 세 리스트를 각각 "카테고리 제목 + 조회 시각 + 순위 / 종목명 / 등락률" 위주의
   별도 텍스트로 정리한다.
   (단, 시가총액상위는 ka10099 응답에 등락률 필드가 없어 대신 시가총액(억원)을 보여준다.
    이는 이전 대화에서 확인한 ka10099 API 자체의 한계다.)
6. send_kakao_message.load_access_token()/build_text_template()/send_memo()를
   그대로 호출해 카테고리별 메시지를 전송 전 글자 수를 출력하며 순차적으로
   카카오톡으로 전송한다.
7. 위 3개 메시지 전송과는 별개로, data/daily_ranking_log.csv를 다시 읽어들여 카테고리별
   (실시간종목조회순위/거래대금상위/시가총액상위)로 시트를 나눈 "3일 비교표" 엑셀 파일을
   만들어 OneDrive(C:\\Users\\pc\\OneDrive\\주식_3일비교.xlsx)에 저장한다
   (save_three_day_comparison_excel()). 시트 1행에는 "{카테고리} 3일 비교표" 제목을 큰
   글씨로 넣어 A1부터 표의 마지막 날짜 열까지 병합·가운데 정렬하고(merge_title_row()),
   2행부터 최근 3개 날짜를 열로, 순위 1~20을 행으로 삼아 "종목명(값)" 형식의 표를 만든다.
   순위 열(A열)은 "1"~"20" 숫자만 들어가므로 RANK_COLUMN_WIDTH(5) 고정 너비를 쓰고,
   날짜 열은 제목 행을 제외한 나머지 셀의 표시 너비(한글 등 동아시아 폭 넓은 문자는 2,
   그 외는 1로 계산)에 최소한의 여유(COLUMN_WIDTH_PADDING=2)만 더해 자동으로 조정한다
   (autosize_columns()). 같은 날짜에 오전/오후 두 번 기록되는 것을 감안해 날짜별로 가장
   나중 기록만 사용하며, 기록된 날짜가 3일 미만이면 있는 날짜만큼만 나열한다. 파일이 이미
   있으면 새로 만들지 않고 열어서 3개 시트의 표 영역 "값"만 최신 CSV로 덮어쓰므로, 사람이
   엑셀에서 직접 추가한 다른 시트/메모/서식은 다음 실행에도 사라지지 않는다(단, 표 영역
   안의 값 자체를 사람이 고쳐도 다음 실행 때 다시 최신 값으로 덮어써진다). 엑셀 파일 저장
   후에는 표 내용을 그대로 옮겨 적지 않고, "3일 비교표가 업데이트됐습니다. OneDrive에서
   확인하세요"라는 짧은 안내 메시지 1개만 카카오톡으로 보낸다. 이 기능은 4~6번 기존 로직을
   전혀 수정하지 않고 별도 함수로만 추가되어 있다.
"""

import csv
import os
import sys
import unicodedata
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# 아래 세 모듈은 오늘 이미 만들어둔 스크립트를 그대로 import해서 재사용한다.
# (모듈 이름 그대로 쓰면 매번 "get_kiwoom_ranking.xxx"처럼 길어지므로 별칭을 붙인다)
import get_kiwoom_ranking as ranking
import get_kiwoom_marketcap_ranking as marketcap
import send_kakao_message as kakao

# Windows 콘솔(cp949) 환경에서 한글/이모지 출력 시 오류가 나는 것을 방지
sys.stdout.reconfigure(encoding="utf-8")

# 각 순위 항목별로 몇 개씩 뽑아 보여줄지
TOP_N = 20

# 실시간종목조회순위/거래대금상위 TOP_N을 누적 기록하는 CSV 로그 경로
LOG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "daily_ranking_log.csv")
LOG_HEADER = ["날짜", "시간", "카테고리", "순위", "종목명", "종목코드", "등락률"]

# 카테고리별 3일 비교표를 저장할 엑셀 파일 경로 (OneDrive에 저장해 다른 기기에서도 확인 가능)
COMPARISON_EXCEL_PATH = r"C:\Users\pc\OneDrive\주식_3일비교.xlsx"


def fetch_inquiry_and_value_top5(token: str) -> dict:
    """
    실시간종목조회순위(ka00198)와 거래대금상위(ka10032)를 조회해 각각 상위 TOP_N개를 돌려준다.

    get_kiwoom_ranking.request_ranking() 함수는 자기 모듈의 전역 변수인
    ranking.TR_ID 값을 읽어서 어떤 TR을 호출할지 결정하도록 만들어져 있다.
    (get_kiwoom_ranking.py를 "python get_kiwoom_ranking.py ka10032"처럼
    커맨드라인 인자로 실행하던 방식과 동일한 동작을 함수 재사용으로 흉내내는 것)
    그래서 여기서는 호출 직전에 ranking.TR_ID 값을 바꿔가며 두 번 호출한다.
    """
    results = {}

    for tr_id in ("ka00198", "ka10032"):
        ranking.TR_ID = tr_id  # request_ranking()이 내부적으로 참조하는 TR 코드를 전환

        tr_name = ranking.TR_NAMES.get(tr_id, "")
        print(f"[INFO] {tr_id}({tr_name}) 조회 중...")
        data = ranking.request_ranking(token)

        # 응답 JSON에서 실제 순위 리스트가 들어있는 키를 찾는다.
        # TR마다 키 이름이 달라서(tdy_trde_qty_upper, trde_prica_upper 등)
        # "리스트 타입이면서 비어있지 않은 값"을 가진 첫 번째 키를 사용한다.
        list_key = next(
            (k for k, v in data.items() if isinstance(v, list) and v),
            None,
        )
        if not list_key:
            print(f"[WARN] {tr_id} 응답에서 리스트 데이터를 찾지 못했습니다: {data}")
            results[tr_id] = []
            continue

        # ETF/ETN, 우선주 제외 (get_kiwoom_ranking.is_common_stock 재사용)
        filtered = [
            item for item in data[list_key]
            if ranking.is_common_stock(item.get("stk_nm", ""))
        ]
        results[tr_id] = filtered[:TOP_N]

    return results


def fetch_marketcap_top5(token: str) -> list:
    """
    KOSPI(mrkt_tp=0)+ KOSDAQ(mrkt_tp=10) 전체 종목을 모아 시가총액 상위 TOP_N개를 돌려준다.

    get_kiwoom_marketcap_ranking.py와 마찬가지로, ka10099는 "순위 조회 TR"이
    아니라 시장별 전체 종목 목록을 주는 TR이기 때문에 직접 시가총액을 계산해서
    정렬해야 한다.
    """
    all_items = []
    for mrkt_tp, label in marketcap.MARKETS.items():
        print(f"[INFO] ka10099(종목정보 리스트) {label}(mrkt_tp={mrkt_tp}) 조회 중...")
        all_items.extend(marketcap.request_stock_list(token, mrkt_tp))

    # ETF/ETN/리츠, 우선주 제외 (get_kiwoom_marketcap_ranking.is_common_stock 재사용)
    common = [item for item in all_items if marketcap.is_common_stock(item)]

    # 시가총액 = 상장주식수(listCount) x 현재가(lastPrice)
    for item in common:
        item["market_cap"] = int(item["listCount"]) * int(item["lastPrice"])

    ranked = sorted(common, key=lambda x: x["market_cap"], reverse=True)
    top = ranked[:TOP_N]

    # CSV 로그에는 등락률 대신 시가총액(억원)을 기록한다 (ka10099에는 등락률 필드가 없음)
    for item in top:
        item["market_cap_eok"] = item["market_cap"] // 100_000_000

    return top


def append_ranking_log(
    category: str,
    items: list,
    rate_field: str,
    date_str: str,
    time_str: str,
    name_field: str = "stk_nm",
    code_field: str = "stk_cd",
) -> None:
    """
    카테고리 하나의 TOP_N 리스트를 data/daily_ranking_log.csv에 이어서(append) 기록한다.

    파일이 없으면 헤더를 포함해 새로 만들고, 있으면 헤더 없이 데이터 행만 추가한다.
    실시간종목조회순위(ka00198)는 등락률 필드명이 base_comp_chgr, 거래대금상위(ka10032)는
    flu_rt로 서로 달라서 어떤 필드를 등락률로 쓸지 rate_field로 지정받는다.
    종목명/종목코드 필드명도 카테고리마다 달라서(ka00198/ka10032는 stk_nm/stk_cd,
    ka10099 기반 시가총액상위는 name/code) name_field/code_field로 지정받는다.
    """
    os.makedirs(os.path.dirname(LOG_PATH), exist_ok=True)
    file_exists = os.path.exists(LOG_PATH)

    with open(LOG_PATH, "a", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(LOG_HEADER)
        for i, item in enumerate(items, start=1):
            writer.writerow([
                date_str,
                time_str,
                category,
                i,
                item.get(name_field, ""),
                item.get(code_field, ""),
                item.get(rate_field, ""),
            ])

    print(f"[INFO] {category} TOP{len(items)}을(를) {LOG_PATH}에 기록했습니다.")


# 카카오 "나에게 보내기" 텍스트 메시지의 안전한 길이 기준(글자 수)
MAX_MESSAGE_LENGTH = 200


def build_inquiry_message(inquiry_top: list, queried_at: str) -> str:
    """
    실시간종목조회순위(ka00198) 리스트를 "[실시간종목조회순위 TOP{TOP_N}]" 제목의
    메시지 본문으로 정리한다.

    ka00198 응답에는 ka10030/ka10032와 달리 flu_rt(등락률) 필드가 없고, 대신
    base_comp_chgr(기준가 대비 등락율, 부호 포함 %)가 온다.
    """
    lines = [f"[실시간종목조회순위 TOP{TOP_N}]", queried_at, ""]
    for i, item in enumerate(inquiry_top, start=1):
        lines.append(f"{i}. {item.get('stk_nm', '')} ({item.get('base_comp_chgr', '')}%)")
    return "\n".join(lines)


def format_rank_change(item: dict) -> str:
    """
    ka10032 응답의 now_rank(현재순위)/pred_rank(전일순위)로 전일 대비 순위 변동을 계산한다.

    ka10032에는 ka00198의 rank_chg처럼 변동값을 바로 주는 필드가 없어서 직접 뺄셈으로 구한다.
    순위가 올랐으면(숫자가 작아졌으면) ▲n, 내렸으면 ▼n, 그대로면 '-'.
    """
    try:
        diff = int(item.get("pred_rank")) - int(item.get("now_rank"))
    except (TypeError, ValueError):
        return "-"

    if diff > 0:
        return f"▲{diff}"
    if diff < 0:
        return f"▼{-diff}"
    return "-"


def build_value_message(value_top: list, queried_at: str) -> str:
    """거래대금상위(ka10032) 리스트를 "[거래대금상위 TOP{TOP_N}]" 제목의 메시지 본문으로 정리한다."""
    lines = [f"[거래대금상위 TOP{TOP_N}]", queried_at, ""]
    for i, item in enumerate(value_top, start=1):
        rank_change = format_rank_change(item)
        lines.append(f"{i}. {item.get('stk_nm', '')} ({item.get('flu_rt', '')}%) [{rank_change}]")
    return "\n".join(lines)


def build_marketcap_message(marketcap_top: list, queried_at: str) -> str:
    """
    시가총액상위(ka10099 기반) 리스트를 "[시가총액상위 TOP{TOP_N}]" 제목의 메시지 본문으로 정리한다.

    ka10099 응답에는 등락률 필드가 없어서, 대신 계산한 시가총액을 억원 단위로 보여준다
    (name 필드, market_cap은 fetch_marketcap_top5에서 계산해둠).
    """
    lines = [f"[시가총액상위 TOP{TOP_N}]", queried_at, ""]
    for i, item in enumerate(marketcap_top, start=1):
        cap_eok = item["market_cap"] // 100_000_000
        lines.append(f"{i}. {item.get('name', '')} ({cap_eok:,}억원)")
    return "\n".join(lines)


def send_category_message(kakao_token: str, title: str, message_text: str) -> None:
    """
    카테고리 하나의 메시지 본문을 글자 수와 함께 출력하고, 카카오톡으로 전송한다.

    카카오 "나에게 보내기" 텍스트 템플릿은 너무 긴 메시지를 보내면 실패할 수 있어서,
    전송 전에 글자 수를 출력해 MAX_MESSAGE_LENGTH(200자) 이내인지 눈으로 확인할 수 있게 한다.
    """
    length = len(message_text)
    status = "OK" if length <= MAX_MESSAGE_LENGTH else "WARN: 200자 초과"
    print(f"[INFO] {title} 메시지 길이: {length}자 ({status})")

    print(f"----- {title} 전송할 메시지 미리보기 -----")
    print(message_text)
    print("----------------------------------\n")

    template_object = kakao.build_text_template(message_text)
    print(f"[INFO] {title} 메시지를 카카오톡 '나에게 보내기'로 전송하는 중입니다...")
    result = kakao.send_memo(kakao_token, template_object)

    if result.get("result_code") == 0:
        print(f"[INFO] {title} 메시지 전송에 성공했습니다.")
    else:
        print(f"[WARN] {title} 메시지 전송 중 예상치 못한 응답을 받았습니다: {result}")


def load_ranking_log_rows() -> list:
    """
    data/daily_ranking_log.csv 전체를 딕셔너리 행 리스트로 읽어 돌려준다.
    (3일치 비교 메시지 전용 함수. 기존 append_ranking_log()/CSV 포맷은 건드리지 않고 읽기만 한다)
    """
    if not os.path.exists(LOG_PATH):
        return []
    with open(LOG_PATH, "r", newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def latest_snapshot_per_date(rows: list, category: str) -> dict:
    """
    카테고리 하나에 대해 날짜별로 그날 가장 나중 시간에 기록된 스냅샷(그 시각에 기록된
    전체 행, 순위 오름차순)을 모아 {날짜: [행, ...]} 형태로 돌려준다.

    하루에 오전/오후 두 번 기록되는 것을 감안해, 같은 날짜에 여러 시간이 있으면
    "HH:MM" 문자열 중 가장 큰(=가장 늦은) 시간의 기록만 사용한다.
    """
    by_date_time = {}
    for row in rows:
        if row.get("카테고리") != category:
            continue
        by_date_time.setdefault(row["날짜"], {}).setdefault(row["시간"], []).append(row)

    snapshots = {}
    for date, by_time in by_date_time.items():
        latest_time = max(by_time.keys())
        snapshots[date] = sorted(by_time[latest_time], key=lambda r: int(r["순위"]))
    return snapshots


def format_comparison_value(category: str, raw_value: str) -> str:
    """
    3일 비교 메시지에서 종목 옆에 괄호로 붙일 값을 카테고리에 맞는 단위로 포맷한다.

    실시간종목조회순위/거래대금상위는 CSV "등락률" 칸에 등락률(%) 값이 들어있고,
    시가총액상위는 같은 칸에 시가총액(억원) 값이 들어있어(append_ranking_log 참고) 단위가 다르다.
    """
    if category == "시가총액상위":
        try:
            return f"{int(raw_value):,}억원"
        except (TypeError, ValueError):
            return f"{raw_value}억원"
    return f"{raw_value}%"


# 날짜 열(종목명+등락률이 들어가는 열) 너비를 셀 내용 표시 너비에 맞출 때 더해줄 여유분.
# 너무 크면 열이 필요 이상으로 넓어지므로, 안 잘리는 선에서 최소한으로만 둔다.
COLUMN_WIDTH_PADDING = 2

# 순위 열(A열, "1"~"20" 숫자만 들어감)은 내용 기준 자동계산 없이 고정 너비를 쓴다.
RANK_COLUMN_INDEX = 1
RANK_COLUMN_WIDTH = 5

# 제목 행(1행)은 열 너비 계산에서 제외한다 (제목이 훨씬 길어서 넣으면 열이 과도하게 넓어짐)
TITLE_ROW = 1


def display_width(text: str) -> int:
    """
    문자열의 "표시 너비"를 센다. 한글/한자 등 동아시아 폭이 넓은(Wide/Fullwidth) 문자는
    영문 폰트 기준 글자 2개 폭을 차지하므로 2로, 나머지(영문/숫자/기호 등)는 1로 센다.
    len()만으로 글자 수를 세면 한글이 많이 섞인 셀(예: "삼성전자(+4.56%)")의 실제 표시
    너비를 과소평가해서 열 너비가 부족해지고 값이 잘려 보이는 문제가 있었다.
    """
    return sum(2 if unicodedata.east_asian_width(ch) in ("W", "F") else 1 for ch in text)


def autosize_columns(ws) -> None:
    """
    openpyxl에는 엑셀의 "열 너비 자동 맞춤" 기능이 없어서, 각 열에서(제목 행 제외) 가장 넓은
    셀 내용의 표시 너비를 직접 재서 그 값 + COLUMN_WIDTH_PADDING을 열 너비로 지정하는
    방식으로 흉내낸다. 다만 순위 열(A열)은 "1"~"20" 숫자만 들어가므로 내용 계산 없이
    RANK_COLUMN_WIDTH 고정값을 쓴다.
    """
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)

        if col_cells[0].column == RANK_COLUMN_INDEX:
            ws.column_dimensions[col_letter].width = RANK_COLUMN_WIDTH
            continue

        max_width = max(
            (display_width(str(cell.value)) for cell in col_cells
             if cell.value is not None and cell.row != TITLE_ROW),
            default=0,
        )
        ws.column_dimensions[col_letter].width = max_width + COLUMN_WIDTH_PADDING


def get_or_create_sheet(workbook: "Workbook", title: str):
    """workbook에 title 이름의 시트가 이미 있으면 그걸 재사용하고, 없으면 새로 만든다."""
    if title in workbook.sheetnames:
        return workbook[title]
    return workbook.create_sheet(title=title)


# 제목 아래로 표를 몇 행 밀어서 시작할지 (1행: 제목, 2행: 표 헤더, 3행부터: 데이터)
TABLE_HEADER_ROW = TITLE_ROW + 1

# 시트 제목 글씨 크기/굵기 (탭 이름과 별개로 시트 안에도 큰 글씨로 보이도록)
TITLE_FONT = Font(size=14, bold=True)
TITLE_ALIGNMENT = Alignment(horizontal="center", vertical="center")

# 순위 열(A열)의 1~20 숫자를 가운데 정렬하기 위한 스타일 (날짜 열의 종목명/등락률 값은 그대로 둔다)
RANK_ALIGNMENT = Alignment(horizontal="center", vertical="center")


def merge_title_row(ws, last_col: int) -> None:
    """
    제목 행(A{TITLE_ROW})을 표의 마지막 열(순위 열 + 날짜 열 전체, last_col)까지 병합하고
    가운데 정렬한다.

    ws는 기존 파일에서 재사용하는 시트일 수 있어서, 이전 실행에서(날짜 열 개수가 달라져)
    병합해 둔 범위가 이번과 다를 수 있다. openpyxl은 겹치는 범위를 다시 병합하려 하면
    오류를 내므로, 제목 행에 걸쳐 있는 기존 병합은 먼저 모두 해제한 뒤 새로 병합한다.
    """
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row <= TITLE_ROW <= merged_range.max_row:
            ws.unmerge_cells(str(merged_range))

    if last_col > 1:
        ws.merge_cells(start_row=TITLE_ROW, start_column=1, end_row=TITLE_ROW, end_column=last_col)

    ws.cell(row=TITLE_ROW, column=1).alignment = TITLE_ALIGNMENT


def fill_comparison_sheet(ws, category: str, rows: list) -> None:
    """
    카테고리 하나에 대해, 1행에는 "{category} 3일 비교표" 제목을 큰 글씨로 넣고 표의 마지막
    열까지 병합·가운데 정렬하며, 그 아래(2행부터)에 최근 3개 날짜(하루 중 가장 나중 기록
    기준)를 열로, 순위 1~TOP_N을 행으로 하는 표를 시트에 채운다. 각 셀은 "종목명(값)"
    형식으로 채운다(값의 단위는 format_comparison_value가 카테고리에 맞게 결정한다).

    ws는 매번 새로 만드는 게 아니라 기존 파일에서 재사용하는 시트일 수 있어서, 표 영역 밖의
    다른 셀(사람이 추가한 메모 등)이나 셀 서식은 건드리지 않고, 제목/표 영역의 "값"만 최신
    CSV 기준으로 덮어쓴다.
    """
    ws.cell(row=TITLE_ROW, column=1, value=f"{category} 3일 비교표").font = TITLE_FONT

    ws.cell(row=TABLE_HEADER_ROW, column=1, value="순위")

    snapshots = latest_snapshot_per_date(rows, category)
    dates = sorted(snapshots.keys())[-3:]  # 오래된 -> 최신 순으로 최근 3개 날짜

    for col, d in enumerate(dates, start=2):
        date_label = datetime.strptime(d, "%Y-%m-%d").strftime("%m/%d")
        ws.cell(row=TABLE_HEADER_ROW, column=col, value=date_label)

    for rank in range(1, TOP_N + 1):
        data_row = TABLE_HEADER_ROW + rank
        ws.cell(row=data_row, column=1, value=rank).alignment = RANK_ALIGNMENT
        for col, d in enumerate(dates, start=2):
            item = next((it for it in snapshots[d] if int(it["순위"]) == rank), None)
            if item is None:
                continue
            name = item.get("종목명", "")
            value = format_comparison_value(category, item.get("등락률", ""))
            ws.cell(row=data_row, column=col, value=f"{name}({value})")

    merge_title_row(ws, last_col=1 + len(dates))
    autosize_columns(ws)


def save_three_day_comparison_excel(rows: list, path: str = COMPARISON_EXCEL_PATH) -> None:
    """
    기존 3개 카카오톡 메시지(TOP20) 전송과는 별개로, 카테고리별로 시트를 나눈 3일 비교표를
    엑셀 파일로 만들어 OneDrive에 저장한다. 기존 조회/전송 로직에는 전혀 관여하지 않고
    CSV 로그를 다시 읽어들이는 것만으로 동작한다.

    파일이 이미 있으면 매번 새로 만들지 않고 load_workbook()으로 열어서 재사용한다.
    사람이 엑셀 파일에서 직접 서식을 바꾸거나 메모를 남기거나 시트를 추가해 두었어도,
    이 함수는 카테고리 3개(실시간종목조회순위/거래대금상위/시가총액상위) 시트의 표 영역
    "값"만 최신 CSV 기준으로 덮어쓸 뿐 그 밖의 시트나 서식은 건드리지 않으므로 사람이 손댄
    내용이 다음 실행 때 사라지지 않는다. 다만 표 영역 안의 셀 값(종목명/등락률 등)은
    CSV의 최신 조회 결과를 그대로 반영해야 하므로, 그 부분을 사람이 직접 고쳐도 다음
    실행 때 다시 최신 값으로 덮어써진다.
    """
    if os.path.exists(path):
        workbook = load_workbook(path)
    else:
        workbook = Workbook()
        workbook.remove(workbook.active)  # openpyxl이 기본으로 만드는 빈 시트를 제거

    for category in ("실시간종목조회순위", "거래대금상위", "시가총액상위"):
        ws = get_or_create_sheet(workbook, category)
        fill_comparison_sheet(ws, category, rows)

    os.makedirs(os.path.dirname(path), exist_ok=True)
    workbook.save(path)
    print(f"[INFO] 3일 비교표를 {path}에 저장했습니다.")


def main():
    # 1. 키움 접근토큰을 한 번만 발급받아 세 번의 조회에 공용으로 사용한다
    app_key, secret_key = ranking.load_app_credentials()
    token = ranking.issue_access_token(app_key, secret_key)

    # 2. 실시간종목조회순위/거래대금상위 조회
    vv = fetch_inquiry_and_value_top5(token)

    # 3. 시가총액상위 조회
    mc_top5 = fetch_marketcap_top5(token)

    # 4. 실시간종목조회순위/거래대금상위 TOP_N을 CSV 로그에 이어서 기록한다
    # (카카오톡 전송 성공 여부와 무관하게, 조회할 때마다 남긴다)
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M")
    append_ranking_log("실시간종목조회순위", vv["ka00198"], "base_comp_chgr", date_str, time_str)
    append_ranking_log("거래대금상위", vv["ka10032"], "flu_rt", date_str, time_str)
    append_ranking_log(
        "시가총액상위", mc_top5, "market_cap_eok", date_str, time_str,
        name_field="name", code_field="code",
    )

    # 5. 카테고리별로 메시지 본문을 따로 정리한다 (하나로 합치지 않는다)
    # 세 카테고리 모두 같은 실행에서 조회한 것이므로 조회 시각도 하나로 통일해서 붙인다
    queried_at = f"{date_str} {time_str}"
    inquiry_message = build_inquiry_message(vv["ka00198"], queried_at)
    value_message = build_value_message(vv["ka10032"], queried_at)
    marketcap_message = build_marketcap_message(mc_top5, queried_at)

    # 6. send_kakao_message.py의 전송 기능을 재사용해 카테고리별로 순차 전송한다
    kakao_token = kakao.load_access_token()

    send_category_message(kakao_token, "실시간종목조회순위", inquiry_message)
    send_category_message(kakao_token, "거래대금상위", value_message)
    send_category_message(kakao_token, "시가총액상위", marketcap_message)

    # 7. 기존 3개 메시지 전송과는 별도로, 카테고리별 3일 비교표 엑셀 파일을 OneDrive에 저장하고
    # (data/daily_ranking_log.csv를 다시 읽어들여 만들며, 위 6번까지의 동작에는 영향을 주지 않는다)
    # 표 내용을 옮겨 적는 대신 업데이트 사실만 짧게 카카오톡으로 알린다
    comparison_rows = load_ranking_log_rows()
    save_three_day_comparison_excel(comparison_rows)
    send_category_message(
        kakao_token, "3일비교표 업데이트 알림",
        "3일 비교표가 업데이트됐습니다. OneDrive에서 확인하세요",
    )


if __name__ == "__main__":
    main()
